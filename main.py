import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import zipfile
import random
from selenium.webdriver.common.keys import Keys
import pandas as pd
import xlwt
import win32com.client
import psutil
import threading
import openpyxl
import xlrd

# 配置文件路径
CONFIG_FILE = 'app_config.json'


class Logger:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def log(self, message, level="INFO"):
        timestamp = time.strftime("[%Y-%m-%d %H:%M:%S]")
        log_message = f"{timestamp} [{level}] {message}\n"
        self.text_widget.insert(tk.END, log_message)
        self.text_widget.see(tk.END)
        self.text_widget.update()


def save_config(excel_path, account, password, driver_path):
    """保存所有配置信息"""
    config = {
        'excel_path': excel_path,
        'account': account,
        'password': password,
        'driver_path': driver_path
    }
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)


def load_config():
    """加载配置信息"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def check_local_chromedriver(logger):
    """检查本地是否存在ChromeDriver"""
    common_paths = [
        "D:\\chromedriver-win64\\chromedriver.exe",
        "D:\\chromedriver_win32\\chromedriver.exe",
        "D:\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe",
        "D:\\chromedriver_win32\\chromedriver-win32\\chromedriver.exe"
    ]

    for path in common_paths:
        if os.path.exists(path):
            logger.log(f"找到本地ChromeDriver: {path}", "SUCCESS")
            return path

    logger.log("未找到本地ChromeDriver", "INFO")
    return None


def get_chrome_version():
    """获取Chrome浏览器版本"""
    chrome_paths = [
        'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe',
        'C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe'
    ]

    for chrome_path in chrome_paths:
        if os.path.exists(chrome_path):
            try:
                from win32com.client import Dispatch
                parser = Dispatch("Scripting.FileSystemObject")
                version = parser.GetFileVersion(chrome_path)
                return version.split('.')[0]
            except Exception:
                continue
    return None


def check_internet_connection(logger):
    """检查网络连接"""
    sites = [
        "https://www.alibaba.com",
        "https://www.taobao.com",
        "https://www.baidu.com"
    ]

    for site in sites:
        try:
            response = requests.get(site, timeout=10)
            if response.status_code == 200:
                logger.log(f"网络连接正常 ({site})", "SUCCESS")
                return True
        except:
            continue

    logger.log("网络连接失败，请检查网络设置", "ERROR")
    return False


def get_chrome_config():
    """获取Chrome配置"""
    options = webdriver.ChromeOptions()

    # 基本设置
    options.add_argument('--no-sandbox')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-software-rasterizer')

    # 性能优化
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-logging')
    options.add_argument('--disable-notifications')
    options.add_argument('--disable-default-apps')

    # 避免检测
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)

    # 设置窗口大小
    options.add_argument('--window-size=1000,800')

    return options


def download_chromedriver_from_official(logger):
    """从Chrome官方下载ChromeDriver"""
    try:
        chrome_version = get_chrome_version()
        if not chrome_version:
            logger.log("无法获取Chrome版本", "ERROR")
            return None

        logger.log(f"检测到Chrome版本: {chrome_version}", "INFO")

        # 准备目录
        driver_dir = os.path.join(os.getcwd(), "chromedriver")
        os.makedirs(driver_dir, exist_ok=True)
        driver_path = os.path.join(driver_dir, "chromedriver.exe")

        # 如果已存在驱动，直接返回
        if os.path.exists(driver_path):
            logger.log(f"使用已存在的ChromeDriver: {driver_path}", "INFO")
            return driver_path

        # 获取体版本号
        logger.log("正在获取最新版本信息...", "INFO")
        version_url = "https://googlechromelabs.github.io/chrome-for-testing/LATEST_RELEASE_" + chrome_version

        try:
            response = requests.get(version_url, timeout=10)
            if not response.ok:
                logger.log("无法获取版本信息", "ERROR")
                return None

            specific_version = response.text.strip()
            logger.log(f"找到匹配版本: {specific_version}", "INFO")

            # 构建下载URL
            download_url = f"https://edgedl.me.gvt1.com/edgedl/chrome/chrome-for-testing/{specific_version}/win64/chromedriver-win64.zip"
            logger.log(f"正在从官方源下载: {download_url}", "INFO")

            # 下载文件
            response = requests.get(download_url, stream=True, timeout=30)
            if not response.ok:
                logger.log("下载失败", "ERROR")
                return None

            # 保存文件
            zip_path = os.path.join(driver_dir, "chromedriver.zip")
            total_size = int(response.headers.get('content-length', 0))
            block_size = 1024

            with open(zip_path, 'wb') as f:
                downloaded = 0
                for data in response.iter_content(block_size):
                    downloaded += len(data)
                    f.write(data)
                    if total_size > 0:
                        percent = int((downloaded / total_size) * 100)
                        if percent % 10 == 0:
                            logger.log(f"下载进度: {percent}%", "INFO")

            logger.log("下载完成，正在解压...", "INFO")

            # 解压文件
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(driver_dir)

            # 移动文件到正确位置
            chromedriver_dir = os.path.join(driver_dir, "chromedriver-win64")
            if os.path.exists(chromedriver_dir):
                source_driver = os.path.join(chromedriver_dir, "chromedriver.exe")
                if os.path.exists(source_driver):
                    import shutil
                    shutil.move(source_driver, driver_path)
                    shutil.rmtree(chromedriver_dir)

            # 删除zip文件
            os.remove(zip_path)

            if os.path.exists(driver_path):
                logger.log(f"ChromeDriver成功: {driver_path}", "SUCCESS")
                return driver_path
            else:
                logger.log("ChromeDriver安装失败", "ERROR")
                return None

        except Exception as e:
            logger.log(f"载过程中发生错误: {str(e)}", "ERROR")
            return None

    except Exception as e:
        logger.log(f"载ChromeDriver发生错误: {str(e)}", "ERROR")
        return None


def get_manual_driver_path(logger):
    """动择ChromeDriver路径"""
    logger.log("请手动选择ChromeDriver文件...", "INFO")
    driver_path = filedialog.askopenfilename(
        title='选择ChromeDriver文件',
        filetypes=[('ChromeDriver', 'chromedriver.exe'), ('All Files', '*.*')]
    )

    if driver_path:
        logger.log(f"已选择ChromeDriver: {driver_path}", "SUCCESS")
        return driver_path
    return None


def get_driver_path(logger):
    """获取ChromeDriver路径"""
    # 首先检查配置文件
    config = load_config()
    driver_path = config.get('driver_path')
    if driver_path and os.path.exists(driver_path):
        logger.log(f"使用已保存的ChromeDriver配置: {driver_path}", "INFO")
        return driver_path

    # 然后检查本地目录
    driver_path = check_local_chromedriver(logger)
    if driver_path:
        # 保存找到的路径到配置文件
        config['driver_path'] = driver_path
        save_config(
            config.get('excel_path', ''),
            config.get('account', ''),
            config.get('password', ''),
            driver_path
        )
        return driver_path

    # 如果本地没有，试自动载
    logger.log("尝试自动下载ChromeDriver...", "INFO")
    driver_path = download_chromedriver_from_official(logger)

    # 如果自动下载失败，提示手动选择
    if not driver_path:
        logger.log("自动下载失败，请手动选择ChromeDriver文件", "WARNING")
        result = messagebox.askyesno("提示",
                                     "未找到ChromeDriver，是否手动选择ChromeDriver文件？\n\n" +
                                     "您可以从以下地址下载对应版本的ChromeDriver：\n" +
                                     "https://googlechromelabs.github.io/chrome-for-testing/\n\n" +
                                     "请确保下载的版本与Chrome浏览器版本匹配。"
                                     )

        if result:
            driver_path = get_manual_driver_path(logger)
            if driver_path:
                # 保存成功选择的路径
                config['driver_path'] = driver_path
                save_config(
                    config.get('excel_path', ''),
                    config.get('account', ''),
                    config.get('password', ''),
                    driver_path
                )
            else:
                logger.log("未选择ChromeDriver文件，操作取消", "WARNING")
                return None
        else:
            logger.log("操作取消", "WARNING")
            return None

    return driver_path


class Application:
    def __init__(self, root):
        self.root = root
        self.root.title("阿里巴巴邮箱数据获取工具")

        # 获取屏幕宽度和高度
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        # 设置窗口大小
        window_width = 800
        window_height = 600

        # 计算窗口位置，使其显示在屏幕右侧
        x_position = screen_width - window_width - 20  # 20是与屏幕右边缘的间距
        y_position = (screen_height - window_height) // 2  # 垂直居中

        # 设置窗口位置和大小
        self.root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")

        # 设置窗口图标
        try:
            icon_path = 'xdlovelife1.ico'
            if os.path.exists(icon_path):
                self.master.iconbitmap(icon_path)
        except Exception as e:
            print(f"设置图标失败: {str(e)}")

        # 禁止调整窗口大小
        self.root.resizable(False, False)

        self.paused = False
        self.running = False
        self.pause_event = threading.Event()
        self.pause_event.set()

        # 创建状态变量
        self.status_var = tk.StringVar(value="就绪")

        # 创建界面
        self.create_widgets()

        # 加载保存的配置
        self.load_saved_config()

        # 底部标签
        self.footer_label = ttk.Label(self.root, text="Powered by xdlovelife", anchor='center')
        self.footer_label.pack(side=tk.BOTTOM, fill=tk.X)

    def load_saved_config(self):
        """加载保存的配置并填充到界面"""
        config = load_config()
        if config:
            # 填Excel路径
            if 'excel_path' in config and os.path.exists(config['excel_path']):
                self.excel_path.set(config['excel_path'])

            # 填充账号
            if 'account' in config:
                self.account_entry.delete(0, tk.END)
                self.account_entry.insert(0, config['account'])

            # 填充密码
            if 'password' in config:
                self.password_entry.delete(0, tk.END)
                self.password_entry.insert(0, config['password'])

    def select_excel(self):
        """选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title='选择Excel文件',
            filetypes=[('Excel Files', '*.xlsx;*.xls'), ('All Files', '*.*')]
        )
        if file_path:
            self.excel_path.set(file_path)
            # 保存配置
            self.save_current_config()

    def save_current_config(self):
        """保存当前配置"""
        config = load_config()
        save_config(
            self.excel_path.get(),
            self.account_entry.get(),
            self.password_entry.get(),
            config.get('driver_path', '')
        )

    def execute(self):
        """执行操作"""
        try:
            if not self.running:
                self.running = True
                self.should_stop = False
                self.paused = False

                # 禁用执行按钮，启用暂停按钮
                self.execute_button.config(state=tk.DISABLED)
                self.pause_button.config(state=tk.NORMAL)

                excel_path = self.excel_path.get()
                account = self.account_entry.get()
                password = self.password_entry.get()

                # 保存当前配置
                self.save_current_config()

                # 在新线程中执行操作
                self.current_thread = threading.Thread(
                    target=self._execute_thread,
                    args=(excel_path, account, password)
                )
                self.current_thread.start()

        except Exception as e:
            self.logger.log(f"执行失败: {str(e)}", "ERROR")
            self.reset_ui()

    def toggle_pause(self):
        """切换暂停状态"""
        if self.running:
            self.paused = not self.paused
            if self.paused:
                self.pause_event.clear()  # 暂停
                self.pause_button.config(text="继续")
                self.status_var.set("已暂停")
                self.logger.log("操作已暂停", "WARNING")
            else:
                self.pause_event.set()  # 继续
                self.pause_button.config(text="暂停")
                self.status_var.set("正在执行")
                self.logger.log("操作已继续", "INFO")

    def _execute_thread(self, excel_path, account, password):
        """在新线程中执行操作"""
        try:
            execute_action(excel_path, account, password, self.logger)
        except Exception as e:
            self.logger.log(f"执行失败: {str(e)}", "ERROR")
        finally:
            self.root.after(0, self.reset_ui)

    def update_status(self):
        """更新状态和进度"""
        if self.current_thread and self.current_thread.is_alive():
            self.root.after(100, self.update_status)
        else:
            self.reset_ui()

    def reset_ui(self):
        """重置UI"""
        self.running = False
        self.paused = False
        self.pause_event.set()
        self.execute_button.config(state=tk.NORMAL)
        self.pause_button.config(state=tk.DISABLED, text="暂停")
        self.status_var.set("就绪")

    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="5")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 添加注意事项公告
        notice_frame = ttk.LabelFrame(main_frame, text="注意事项", padding=5)
        notice_frame.pack(fill=tk.X, padx=5, pady=5)

        notices = [
            "1. 请确保Excel文件格式正确，D列为司名称，P列为邮箱地址",
            "2. 程序会自动跳过D列为空或P列已有数据的行",
            "3. 处理过程中请勿关闭浏览器窗口",
            "4. 如需暂停，请点击暂停按钮，避免直接关闭程序",
            "5. 处理完成后数据会自动保存到Excel文件中",
            "6. 如遇到异常，请查看日志信息进行排查"
        ]

        for notice in notices:
            notice_label = ttk.Label(notice_frame, text=notice, wraplength=600)
            notice_label.pack(anchor=tk.W, padx=5, pady=2)

        # 配置框
        config_frame = ttk.LabelFrame(main_frame, text="配置", padding=5)
        config_frame.pack(fill=tk.X, padx=5, pady=5)

        # Excel文件选择
        excel_frame = ttk.Frame(config_frame)
        excel_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(excel_frame, text="Excel文件:").pack(side=tk.LEFT)
        self.excel_path = tk.StringVar()
        ttk.Entry(excel_frame, textvariable=self.excel_path, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(excel_frame, text="选择文件", command=self.select_excel).pack(side=tk.LEFT)

        # 账号密码输入框架
        login_frame = ttk.Frame(config_frame)
        login_frame.pack(fill=tk.X, padx=5, pady=5)

        # 账号输入
        ttk.Label(login_frame, text="账号:").pack(side=tk.LEFT)
        self.account_entry = ttk.Entry(login_frame, width=20)
        self.account_entry.pack(side=tk.LEFT, padx=5)

        # 密码输入
        ttk.Label(login_frame, text="密码:").pack(side=tk.LEFT, padx=(10, 0))
        self.password_entry = ttk.Entry(login_frame, width=20, show="*")
        self.password_entry.pack(side=tk.LEFT, padx=5)

        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)

        # 执行按钮
        self.execute_button = ttk.Button(
            button_frame,
            text="开始执行",
            command=self.execute
        )
        self.execute_button.pack(side=tk.LEFT, padx=5)

        # 暂停按钮
        self.pause_button = ttk.Button(
            button_frame,
            text="暂停",
            command=self.toggle_pause,
            state=tk.DISABLED
        )
        self.pause_button.pack(side=tk.LEFT, padx=5)

        # 日志框架
        log_frame = ttk.LabelFrame(main_frame, text="运行日志", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 日志文本框
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            height=15,
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 初始化日志器
        self.logger = Logger(self.log_text)

        # 状态标签框架
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=5)

        # 状态标签
        self.status_label = ttk.Label(
            status_frame,
            textvariable=self.status_var,
            width=40
        )
        self.status_label.pack(side=tk.LEFT, padx=5)


def wait_for_manual_verification(driver, logger, timeout=300):
    """等待户手动完成验"""
    logger.log("检到验证码，请手动完成验证...", "WARNING")
    messagebox.showinfo("提示", "请手动完验证码，完成后程序将自动继续")

    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            # 检查是已经通过验证（URL再包含login）
            if "login" not in driver.current_url.lower():
                logger.log("验证通过，继续执行...", "SUCCESS")
                return True
            time.sleep(2)
        except:
            pass

    logger.log("验证码等待时", "ERROR")
    return False


def handle_login(driver, account, password, logger):
    """处理登录程"""
    try:
        # 等待用名输入框
        logger.log("等待用户名输入框...")
        username_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[2]/div[2]/input'))
        )
        logger.log("成功找到用户名输入框", "SUCCESS")

        # 输入账号
        logger.log("正在输入账号...")
        type_like_human(username_input, account, logger)
        time.sleep(1)

        # 等待密码输入框
        logger.log("等待密码输入框...")
        password_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[2]/input'))
        )
        logger.log("成功找到密码输入框", "SUCCESS")

        # 输入密码
        logger.log("正在输入密码...")
        type_like_human(password_input, password, logger)
        time.sleep(1)

        # 等待登录按钮
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[2]/button'))
        )
        logger.log("成功找到登录按钮", "SUCCESS")

        # 点击登录按钮
        logger.log("正在点击录按钮...")
        login_button.click()
        logger.log("已点击登录按钮", "SUCCESS")

        # 等待验证码模块
        time.sleep(2)

        # 检查验证码模块
        logger.log("开始检查验证码模块...", "INFO")
        try:
            # 检查验证码容器
            logger.log("正在查找验证码容器...", "INFO")
            captcha_container = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[2]'))
            )

            if captcha_container.is_displayed():
                logger.log("验证码容器可见", "INFO")
                current_url = driver.current_url

                # 直接显示提示，不再检查具体的验证码元素
                logger.log("请完成验证码验证", "WARNING")
                messagebox.showinfo(
                    "需要验证",
                    "请手动完成验证\n\n" +
                    "1. 请滑动滑块完成验证\n" +
                    "2. 验证通过后程序将自动继续\n" +
                    "3. 如果失败请重新滑动"
                )

                # 等待验证完成和页面跳转
                logger.log("待验证完成和页面转...", "INFO")
                max_wait = 300  # 最等待5分钟
                wait_start = time.time()

                while time.time() - wait_start < max_wait:
                    # 检查URL是否改变（登录成功）
                    if driver.current_url != current_url:
                        logger.log("验证通过，检测到页面跳转", "SUCCESS")
                        time.sleep(2)  # 等待页面稳定
                        return True
                    time.sleep(1)

                logger.log("验证等待超时", "ERROR")
                return False

        except Exception as e:
            logger.log(f"验证码检测过程出错: {str(e)}", "WARNING")

        # 检查最终状态
        if "login" not in driver.current_url.lower():
            logger.log("登录成功", "SUCCESS")
            return True
        else:
            logger.log("登录失败，仍在登录页面", "ERROR")
            return False

    except Exception as e:
        logger.log(f"登录过程中发生错误: {str(e)}", "ERROR")
        return False


def handle_post_login(driver, excel_path, logger):
    """处理登录后的操作"""
    try:
        # 等待页面完全加载
        logger.log("等待页面元素加载完成...")
        WebDriverWait(driver, 30).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        logger.log("页面加载完成", "SUCCESS")

        # 给页面一点时间稳定
        time.sleep(3)

        # 记录当前URL用于调试
        current_url = driver.current_url
        logger.log(f"当前页面URL: {current_url}", "INFO")

        # 点击商机沟通菜单
        logger.log("尝试点击商机沟通菜单...", "INFO")
        if not click_business_communication(driver, logger):
            raise Exception("无法进入商机沟通页面")

        # 处理Excel数据
        logger.log("开始处理Excel数据...", "INFO")
        process_excel_data(driver, excel_path, logger)

        logger.log("登录后处理完成", "SUCCESS")
        return True

    except Exception as e:
        logger.log(f"登录后操作失败: {str(e)}", "ERROR")
        return False


def navigate_to_search(driver, logger):
    """导航到搜索页面准备搜索"""
    try:
        logger.log("等待搜索区域加载...")
        wait = WebDriverWait(driver, 20)

        # 点击搜索区域
        logger.log("在点击搜索区域...")
        search_area = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="widget-27"]/div/form/input'))
        )
        search_area.click()
        time.sleep(1)  # 短暂等待

        # 定位发件人输入框
        logger.log("正在定位发件人输入框...")
        sender_input = wait.until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="search-form-sender"]'))
        )

        # 确保输入框可见且可交互
        if sender_input.is_displayed() and sender_input.is_enabled():
            logger.log("成功找到件输入框", "SUCCESS")
            return sender_input
        else:
            raise Exception("发件人输入框不可用")

    except Exception as e:
        logger.log(f"准备搜索时发生错误: {str(e)}", "ERROR")
        return None


def click_business_communication(driver, logger):
    """点击商机沟通菜单"""
    try:
        logger.log("正在等待商机沟通菜单加载...")
        wait = WebDriverWait(driver, 3)

        # 使用JavaScript移除可能的遮罩层
        try:
            driver.execute_script("""
                var elements = document.querySelectorAll('div[class*="modal"], div[class*="dialog"], div[class*="popup"], div[style*="z-index"]');
                elements.forEach(function(element) {
                    element.remove();
                });
            """)
            logger.log("已尝试移除遮罩层", "INFO")
        except:
            pass

        # 然后尝试点击商机通菜单
        menu_element = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="seller-menu-container"]/div/div/div/div[2]/div/ul/div[7]/a'))
        )

        # 使用JavaScript点击，避免被遮挡
        logger.log("正在点击商机沟通菜单...")
        driver.execute_script("arguments[0].click();", menu_element)

        # 等待页面加载完成
        logger.log("等待页面加载完成...")
        wait.until(lambda d: d.execute_script('return document.readyState') == 'complete')
        time.sleep(3)  # 额外等待以确保页面全加载

        # 准备搜索
        sender_input = navigate_to_search(driver, logger)
        if not sender_input:
            raise Exception("无法找到发件人输入框")

        logger.log("商机沟通页面准备完成", "SUCCESS")
        return True

    except Exception as e:
        logger.log(f"击商机沟通菜单时发生错误: {str(e)}", "ERROR")
        return False


def is_file_locked(file_path):
    """检查文件是否被占用"""
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return False

        # 尝试以写入模式打开文件
        try:
            with open(file_path, 'rb+') as f:
                return False
        except (IOError, PermissionError):
            return True

    except Exception:
        return True


def close_excel_file(file_path, logger):
    """尝试正常关闭Excel文件"""
    try:
        # 获取Excel应用程序实例
        excel = win32com.client.GetObject(Class="Excel.Application")

        # 遍历所有打开的工作簿
        for wb in excel.Workbooks:
            try:
                if os.path.abspath(wb.FullName).lower() == os.path.abspath(file_path).lower():
                    wb.Close(SaveChanges=False)
                    logger.log(f"已关闭Excel文件: {file_path}", "SUCCESS")
                    return True
            except:
                continue

        return False
    except:
        return False


def kill_excel_process(logger):
    """关闭所有Excel进程"""
    try:
        killed = False
        # 先尝试使用taskkill命令
        try:
            os.system('taskkill /F /IM excel.exe')
            killed = True
        except:
            pass

        # 如果taskkill失败，使用psutil
        if not killed:
            for proc in psutil.process_iter(['pid', 'name']):
                try:
                    proc_name = proc.info['name'].lower()
                    if 'excel' in proc_name or 'et.exe' in proc_name:  # 添加对WPS的支持
                        proc.kill()
                        killed = True
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    continue

        if killed:
            logger.log("已关闭Excel进程", "SUCCESS")
            time.sleep(1)  # 等待进程完全关闭
            return True
        else:
            logger.log("未找到Excel进程", "INFO")
            return False
    except Exception as e:
        logger.log(f"关闭Excel进程时发生错误: {str(e)}", "ERROR")
        return False


def show_file_locked_dialog(file_path, logger):
    """显示文件被占用时的对话框"""
    response = messagebox.askyesnocancel(
        "文件被占用",
        f"文件 {file_path} 被占用。\n是：关闭Excel进程\n否：保存为临时文件\n取消：放弃保存",
        icon='warning'
    )

    if response is True:
        return "close"
    elif response is False:
        return "temp"
    else:
        return "cancel"


def save_excel_data(df, excel_path, logger):
    """保存Excel数据，处理文件被占用的情况"""
    try:
        # 检查文件类型
        is_xls = excel_path.lower().endswith('.xls')

        # 检查文件是否被占用
        if is_file_locked(excel_path):
            logger.log("Excel文件正在被其他程序使用", "WARNING")

            # 显示文件被占用对话框
            response = show_file_locked_dialog(excel_path, logger)

            if response == "close":
                # 尝试关闭Excel进程
                if kill_excel_process(logger):
                    time.sleep(2)  # 等待进程完全关闭

                    # 再次尝试保存
                    try:
                        if is_xls:
                            # 使用 xlwt 保存
                            wb = xlwt.Workbook(encoding='utf-8')
                            ws = wb.add_sheet('Sheet1')

                            # 写入列名和数据
                            for col_idx, col_name in enumerate(df.columns):
                                ws.write(0, col_idx, str(col_name))

                            for row_idx in range(len(df)):
                                for col_idx in range(len(df.columns)):
                                    try:
                                        value = df.iloc[row_idx, col_idx]
                                        if pd.isna(value):
                                            value = ''
                                        elif isinstance(value, (int, float)):
                                            ws.write(row_idx + 1, col_idx, value)
                                            continue
                                        ws.write(row_idx + 1, col_idx, str(value))
                                    except Exception as cell_error:
                                        logger.log(f"写入单元格 ({row_idx}, {col_idx}) 时发生错误: {str(cell_error)}",
                                                   "WARNING")
                                        ws.write(row_idx + 1, col_idx, '')

                            wb.save(excel_path)
                        else:
                            df.to_excel(excel_path, index=False, engine='openpyxl')

                        logger.log(f"成功保存数据到: {excel_path}", "SUCCESS")
                        return True
                    except Exception as save_error:
                        logger.log(f"保存失败: {str(save_error)}", "ERROR")

            elif response == "temp":
                # 创建临时文件
                temp_path = excel_path.rsplit('.', 1)[0] + '_temp.' + excel_path.rsplit('.', 1)[1]
                try:
                    if is_xls:
                        wb = xlwt.Workbook(encoding='utf-8')
                        ws = wb.add_sheet('Sheet1')
                        # 写入列名和数据
                        for col_idx, col_name in enumerate(df.columns):
                            ws.write(0, col_idx, str(col_name))

                        for row_idx in range(len(df)):
                            for col_idx in range(len(df.columns)):
                                try:
                                    value = df.iloc[row_idx, col_idx]
                                    if pd.isna(value):
                                        value = ''
                                    elif isinstance(value, (int, float)):
                                        ws.write(row_idx + 1, col_idx, value)
                                        continue
                                    ws.write(row_idx + 1, col_idx, str(value))
                                except Exception as cell_error:
                                    logger.log(f"写入单元格 ({row_idx}, {col_idx}) 时发生错误: {str(cell_error)}",
                                               "WARNING")
                                    ws.write(row_idx + 1, col_idx, '')
                        wb.save(temp_path)
                    else:
                        df.to_excel(temp_path, index=False, engine='openpyxl')
                    logger.log(f"已保存到临时文件: {temp_path}", "SUCCESS")
                    return True
                except Exception as temp_error:
                    logger.log(f"保存临时文件失败: {str(temp_error)}", "ERROR")
                    return False
            else:
                logger.log("取消保存", "WARNING")
                return False

        # 如果文件没有被占用，直接保存
        try:
            if is_xls:
                wb = xlwt.Workbook(encoding='utf-8')
                ws = wb.add_sheet('Sheet1')
                # 写入列名和数据
                for col_idx, col_name in enumerate(df.columns):
                    ws.write(0, col_idx, str(col_name))

                for row_idx in range(len(df)):
                    for col_idx in range(len(df.columns)):
                        try:
                            value = df.iloc[row_idx, col_idx]
                            if pd.isna(value):
                                value = ''
                            elif isinstance(value, (int, float)):
                                ws.write(row_idx + 1, col_idx, value)
                                continue
                            ws.write(row_idx + 1, col_idx, str(value))
                        except Exception as cell_error:
                            logger.log(f"写入单元格 ({row_idx}, {col_idx}) 时发生错误: {str(cell_error)}", "WARNING")
                            ws.write(row_idx + 1, col_idx, '')
                wb.save(excel_path)
            else:
                df.to_excel(excel_path, index=False, engine='openpyxl')
            logger.log(f"成功保存数据到: {excel_path}", "SUCCESS")
            return True

        except Exception as save_error:
            logger.log(f"保存文件时发生错误: {str(save_error)}", "ERROR")
            return False

    except Exception as e:
        logger.log(f"保存Excel数据时发生错误: {str(e)}", "ERROR")
        return False


def process_excel_data(driver, excel_path, logger):
    """处理Excel数据"""
    try:
        logger.log("正在检查Excel文件状态...")

        # 检查文件是否被占用
        if is_file_locked(excel_path):
            logger.log("Excel文件正在被其他程序使用", "WARNING")

            # 显示文件被占用对话框
            response = show_file_locked_dialog(excel_path, logger)

            if response == "close":
                # 尝试正常关闭Excel文件
                if not close_excel_file(excel_path, logger):
                    # 如果正常关闭失败，则强制结束Excel进程
                    if not kill_excel_process(logger):
                        logger.log("无法关闭Excel文件，请手动关闭后重试", "ERROR")
                        return False

                # 等待文件释放
                time.sleep(2)

                # 再次检查文件是否被占用
                if is_file_locked(excel_path):
                    logger.log("Excel文件仍然被占用，请手动关闭后重试", "ERROR")
                    return False

            elif response == "temp":
                # 创建临时文件路径
                temp_path = excel_path.rsplit('.', 1)[0] + '_temp.' + excel_path.rsplit('.', 1)[1]
                logger.log(f"将使用临时文件: {temp_path}", "INFO")
                excel_path = temp_path
            else:
                logger.log("操作已取消", "WARNING")
                return False

        logger.log("正在读取Excel文件...")

        # 根据文件扩展名选择不同的读取方式
        if excel_path.lower().endswith('.xls'):
            df = pd.read_excel(excel_path, engine='xlrd')
        else:
            df = pd.read_excel(excel_path, engine='openpyxl')

        num_columns = len(df.columns)
        logger.log(f"Excel文件共有 {num_columns} 列", "INFO")

        # 寻找第一个空白列
        p_column_index = None
        for col_idx in range(num_columns):
            # 检查该列是否为空
            if df.iloc[:, col_idx].isna().all():  # 如果整列都是空的
                p_column_index = col_idx
                logger.log(f"找到空白列，索引为: {col_idx} (第{col_idx + 1}列)", "INFO")
                break

        if p_column_index is None:
            logger.log("未找到空白列，将在文件末尾添加新列", "WARNING")
            # 添加新列
            df[f'Column{num_columns + 1}'] = pd.NA
            p_column_index = num_columns

        logger.log("成功读取Excel文件", "SUCCESS")
        logger.log(f"总行数: {len(df)}", "INFO")

        # 定位发件人输入框
        try:
            sender_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="search-form-sender"]'))
            )
            logger.log("成功找到发件人输入框", "SUCCESS")
        except Exception as e:
            logger.log(f"无法找到发件人输入框: {str(e)}", "ERROR")
            return False

        start_row = 6  # Excel中的第7行
        processed_count = 0
        skipped_count = 0

        while start_row < len(df):
            try:
                # 获取当前P列的值并检查是否为空
                p_column_value = df.iloc[start_row, p_column_index]

                if pd.isna(p_column_value):  # 如果是空值，继续处理
                    # 获取D列的值（第4列，索引3）
                    d_column_index = 3
                    name = str(df.iloc[start_row, d_column_index]).strip()
                    if pd.isna(name) or name == '' or name == 'nan':
                        logger.log(f"跳过第{start_row + 1}行：D列为空", "INFO")
                        skipped_count += 1
                        start_row += 1
                        continue

                    logger.log(f"正在处理第{start_row + 1}行: {name}")

                    try:
                        # 每次操作前重新获取输入框
                        sender_input = WebDriverWait(driver, 20).until(
                            EC.presence_of_element_located((By.XPATH, '//*[@id="search-form-sender"]'))
                        )

                        if sender_input.is_displayed() and sender_input.is_enabled():
                            sender_input.clear()
                            sender_input.send_keys(Keys.CONTROL + "a")
                            sender_input.send_keys(Keys.DELETE)
                            sender_input.send_keys(name)
                            logger.log(f"已输入搜索内容: {name}", "SUCCESS")

                            # 点击搜索按钮
                            search_button = WebDriverWait(driver, 20).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="widget-41"]/form/div[4]/button[1]'))
                            )
                            search_button.click()

                            # 等待搜索结果加载
                            logger.log("等待搜索结果加载...")
                            time.sleep(2)  # 增加等待时间

                            try:
                                # 点击链接打开新窗口
                                link = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable((By.XPATH, '//*[@id="widget-29"]/div[1]'))
                                )
                                link.click()
                                logger.log("已点击搜索结果链接", "INFO")

                                # 切换到新窗口
                                time.sleep(1)  # 等待新窗口打开
                                new_window = driver.window_handles[-1]
                                driver.switch_to.window(new_window)

                                # 等待新页面中的元素加载
                                content_element = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, '//*[@id="app-crm"]/div/div/div[2]/div/div[2]/div/div[2]/div[2]'))
                                )

                                # 获取内容
                                content = content_element.text
                                logger.log(f"获取到内容: {content}")

                                # 关闭新窗口
                                driver.close()

                                # 切回主窗口
                                driver.switch_to.window(driver.window_handles[0])

                                # 写入内容并保存
                                df.iloc[start_row, p_column_index] = content
                                if not save_excel_data(df, excel_path, logger):
                                    logger.log("保存失败，创建备份...", "WARNING")
                                    backup_path = excel_path.rsplit('.', 1)[0] + '_backup.' + excel_path.rsplit('.', 1)[
                                        1]
                                    save_excel_data(df, backup_path, logger)

                                logger.log(f"已将内容保存到第{start_row + 1}行第{p_column_index + 1}列", "SUCCESS")
                                processed_count += 1

                            except Exception as e:
                                logger.log(f"处理新窗口时发生错误: {str(e)}", "ERROR")
                                # 确保回到主窗口
                                try:
                                    if len(driver.window_handles) > 1:
                                        driver.close()
                                    driver.switch_to.window(driver.window_handles[0])
                                except:
                                    pass

                        else:
                            logger.log("发件人输入框不可用，无法输入内容", "ERROR")

                    except Exception as e:
                        logger.log(f"处理第{start_row + 1}行时发生错误: {str(e)}", "ERROR")
                        # 尝试恢复到主窗口
                        try:
                            if len(driver.window_handles) > 1:
                                driver.switch_to.window(driver.window_handles[0])
                        except:
                            pass
                else:
                    logger.log(f"跳过第{start_row + 1}行：目标列已有数据", "INFO")
                    skipped_count += 1

                start_row += 1
                time.sleep(random.uniform(2.0, 3.0))  # 增加随机等待时间

            except Exception as e:
                logger.log(f"处理行时发生错误: {str(e)}", "ERROR")
                start_row += 1
                continue

        logger.log("================================================================", "SUCCESS")
        logger.log("                      处理完成！", "SUCCESS")
        logger.log("----------------------------------------------------------------", "SUCCESS")
        logger.log(f"总共处理: {processed_count}行", "SUCCESS")
        logger.log(f"总共跳过: {skipped_count}行", "SUCCESS")
        logger.log(f"总行数: {len(df)}行", "SUCCESS")
        logger.log("================================================================", "SUCCESS")

        return True

    except Exception as e:
        logger.log(f"处理Excel数据时发生错误: {str(e)}", "ERROR")
        return False


def execute_action(excel_path, account, password, logger):
    """执行主要操作"""
    driver = None
    try:
        if not all([excel_path, account, password]):
            logger.log("错误：请填写所有必要信息！", "ERROR")
            return False

        logger.log("正在检查配置...")
        if not check_paths(logger):
            return False

        # 检查网络连接
        logger.log("正在检查网络连接...")
        if not check_internet_connection(logger):
            messagebox.showerror("错误", "网络连接失败，请检查网络设置！")
            return False

        # 获取Chrome配置和driver
        chrome_options = get_chrome_config()
        driver_path = get_driver_path(logger)
        if not driver_path:
            return False

        try:
            logger.log("正在启动Chrome浏览器...")
            service = Service(driver_path)
            driver = webdriver.Chrome(service=service, options=chrome_options)

            # 设置超时时间
            driver.set_page_load_timeout(60)
            driver.implicitly_wait(20)

            # 直接访问登录页面
            logger.log("正在打开登录页...")
            target_url = "https://i.alibaba.com/index.htm"
            driver.get(target_url)

            # 处理登录
            if not handle_login(driver, account, password, logger):
                raise Exception("登录失败")

            # 登录成功后，等待一下再处理后续操作
            time.sleep(3)

            # 处理登录后的操作
            if not handle_post_login(driver, excel_path, logger):
                logger.log("登录后处理失败", "ERROR")
                return False

            logger.log("所有操作完成", "SUCCESS")
            return True

        except Exception as e:
            error_msg = f"浏览器操作发生错误：{str(e)}"
            logger.log(error_msg, "ERROR")
            messagebox.showerror("错误", error_msg)
            return False

        finally:
            try:
                if driver:
                    driver.quit()
            except:
                pass

    except Exception as e:
        error_msg = f"执行过程中发生错误：{str(e)}"
        logger.log(error_msg, "ERROR")
        messagebox.showerror("错误", error_msg)
        return False


def check_paths(logger):
    """检查必要的径是否存在"""
    chrome_path = 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe'

    if not os.path.exists(chrome_path):
        error_msg = f"Chrome浏览器路径不存在: {chrome_path}\n请检查Chrome是否正确安装！"
        logger.log(error_msg, "ERROR")
        messagebox.showerror("错误", error_msg)
        return False

    logger.log("所有路径检查通过", "SUCCESS")
    return True


def type_like_human(element, text, logger):
    """快速输入文本"""
    try:
        # 清空输入框
        element.clear()
        element.send_keys(Keys.CONTROL + "a")
        element.send_keys(Keys.DELETE)

        # 直接输入文本
        element.send_keys(text)
        return True
    except Exception as e:
        logger.log(f"输入文本时发生错误: {str(e)}", "ERROR")
        return False


def init_driver(logger):
    """初始化浏览器"""
    try:
        # 获取Chrome配置
        options = get_chrome_config()

        # 创建WebDriver实例
        driver = webdriver.Chrome(options=options)

        # 设置窗口位置
        screen_width = driver.execute_script('return window.screen.width')
        x = screen_width - 1020
        driver.set_window_position(x, 20)

        # 置超时
        driver.set_page_load_timeout(30)
        driver.implicitly_wait(10)

        logger.log("浏览器初始化成功", "SUCCESS")
        return driver

    except Exception as e:
        logger.log(f"浏览初始化失败: {str(e)}", "ERROR")
        raise


def get_content_from_new_window(driver, logger):
    """从新窗口获取内容"""
    try:
        # 点击链接打开新窗口
        logger.log("等待搜索结果链接...", "INFO")
        link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="widget-29"]/div[1]'))
        )
        link.click()
        logger.log("已点击搜索结果链接", "INFO")

        # 获取所有窗口句柄
        time.sleep(1)  # 等待新窗口打开
        handles = driver.window_handles

        if len(handles) < 2:
            logger.log("未能打开新窗口", "ERROR")
            return False

        # 切换到新窗口
        new_window = handles[-1]
        driver.switch_to.window(new_window)
        logger.log("已切换到新口", "INFO")

        try:
            # 等待内容加载
            content_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="app-crm"]/div/div/div[2]/div/div[2]/div/div[2]/div[2]'))
            )

            # 获取内容
            content = content_element.text
            logger.log(f"成功获取内容: {content[:50]}...", "SUCCESS")  # 只显示前50个字符

        except Exception as e:
            logger.log(f"获取内容时发生错误: {str(e)}", "ERROR")
            content = None

        finally:
            # 关闭新窗口
            try:
                driver.close()
                logger.log("已关闭新窗口", "INFO")
            except:
                logger.log("关闭新窗口失败", "WARNING")

            # 切回主窗口
            driver.switch_to.window(handles[0])
            logger.log("已切回主窗口", "INFO")

        return content

    except Exception as e:
        logger.log(f"处理新窗口时发生错误: {str(e)}", "ERROR")

        # 尝试恢复到主窗口
        try:
            if len(driver.window_handles) > 1:
                driver.close()  # 关闭当前窗口
            driver.switch_to.window(driver.window_handles[0])  # 切回主窗口
            logger.log("已恢复到主窗口", "INFO")
        except:
            logger.log("恢复主窗口失败", "ERROR")

        return None

    finally:
        # 确保回到主窗口
        try:
            if driver.current_window_handle != driver.window_handles[0]:
                driver.switch_to.window(driver.window_handles[0])
        except:
            pass


def main():
    root = tk.Tk()
    # 设置窗口图标
    try:
        icon_path = 'xdlovelife1.ico'
        if os.path.exists(icon_path):
            root.iconbitmap(icon_path)
    except Exception as e:
        print(f"设置图标失败: {str(e)}")
        
    app = Application(root)
    root.mainloop()


if __name__ == "__main__":
    main()